import openpyxl
import os


class HomeData:
    PAGE_URL = "https://rahulshettyacademy.com/angularpractice/"
    SUCCESS_TEXT_EXPECTED = "Success!"

    @staticmethod
    def get_excel_sheet(filename, sheetname):
        current_dir = os.path.dirname(__file__)
        workbook = openpyxl.load_workbook(current_dir + '\\' + filename)
        return workbook[sheetname]

    @staticmethod
    def get_data(sheet, data_row):
        ''' print(f"\nRetrieving data from row {data_row} ")'''
        if sheet is None:
            sheet = HomeData.get_excel_sheet()

        data = {}
        for col in range(1, sheet.max_column+1):
            data[sheet.cell(1, col).value] = sheet.cell(data_row, col).value
        return data

    @staticmethod
    def get_list_all_data(filename="home_data.xlsx", sheetname="Sheet1"):
        sheet = HomeData.get_excel_sheet(filename, sheetname)
        data_list = []

        for row in range(2, sheet.max_row+1):
            data_list.append(HomeData.get_data(sheet, row))
        return data_list


''' This is for testing the static methods '''
if __name__ == "__main__":
    test = HomeData.get_list_all_data()

